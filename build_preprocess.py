import gc
import json
import math
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "build"
OUT.mkdir(exist_ok=True)

FY_FILE = ROOT / "Opr MIS Data Apr25 to Mar26.xlsx"
NEW_FILE = ROOT / "Opr MIS Data Apr26 to Jul26.xlsx"

HOUSE_CODES = {"A11", "A10287", "A111111", "ERROR"}
GROUPS = ["MSFL", "MSFL-Sharing", "Arbitrage", "P-Sec"]


def clean_text(v):
    if pd.isna(v) or v is None:
        return ""
    return str(v).strip()


def clean_num(s):
    return pd.to_numeric(s, errors="coerce").fillna(0.0)


def read_sheet_fast(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()
    header = [str(c).replace("\n", " ").strip() if c is not None else "" for c in raw_header]
    data = [r for r in rows_iter if any(x is not None for x in r)]
    df = pd.DataFrame(data, columns=header)
    return df


def read_source(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ch = read_sheet_fast(wb, "Channelwise")
    cl = read_sheet_fast(wb, "Clientwise")
    wb.close()
    gc.collect()

    ch.columns = [str(c).replace("\n", " ").strip() for c in ch.columns]
    cl.columns = [str(c).replace("\n", " ").strip() for c in cl.columns]

    ch = ch.rename(columns={
        "Branch Code": "Branch Code",
        "FIRST DATE OF TRADING": "First Date of Trading",
        "Channel ID": "Channel ID",
        "Channel Type": "Channel Type",
        "Group": "Group",
        "Channel Name": "Channel Name",
        "City": "City",
        "STATE": "State",
        "Turnover": "Turnover",
        "Gross Brokerage": "Gross Brokerage",
        "Passout": "Passout",
        "Net Brokerage": "Net Brokerage",
        "Yield": "Yield",
    })
    cl = cl.rename(columns={
        "Branch Code": "Branch Code",
        "Branch Name": "Branch Name",
        "Client Code": "Client Code",
        "Client Name": "Client Name",
        "City": "City",
        "Turnover": "Turnover",
        "Gross Brokerage": "Gross Brokerage",
        "Yield": "Yield",
    })

    for c in ["Branch Code", "First Date of Trading", "Channel ID", "Channel Type", "Group", "Channel Name", "City", "State"]:
        ch[c] = ch[c].map(clean_text)
    for c in ["Branch Code", "Branch Name", "Client Code", "Client Name", "City"]:
        cl[c] = cl[c].map(clean_text)
    for c in ["Turnover", "Gross Brokerage", "Passout", "Net Brokerage", "Yield"]:
        ch[c] = clean_num(ch[c])
    for c in ["Turnover", "Gross Brokerage", "Yield"]:
        cl[c] = clean_num(cl[c])

    ch = ch[ch["Channel ID"] != ""].copy()
    cl = cl[cl["Client Code"] != ""].copy()

    # Aggregate duplicate channel keys if present, preserving a single descriptive row.
    ch_key = ["Branch Code", "Channel ID"]
    ch_meta = ch.sort_values(["Branch Code", "Channel ID"]).groupby(ch_key, as_index=False).first()
    ch_sum = ch.groupby(ch_key, as_index=False)[["Turnover", "Gross Brokerage", "Passout", "Net Brokerage"]].sum()
    ch = ch_meta.drop(columns=["Turnover", "Gross Brokerage", "Passout", "Net Brokerage", "Yield"], errors="ignore").merge(ch_sum, on=ch_key, how="left")
    ch["Yield"] = (ch["Gross Brokerage"] / ch["Turnover"].replace(0, math.nan) * 100).fillna(0)

    # Source-level reconciliation uses the non-total rows from each source.
    source_ch = {
        "turnover": float(ch["Turnover"].sum()),
        "gross": float(ch["Gross Brokerage"].sum()),
        "passout": float(ch["Passout"].sum()),
        "net": float(ch["Net Brokerage"].sum()),
    }
    source_cl = {
        "turnover": float(cl["Turnover"].sum()),
        "gross": float(cl["Gross Brokerage"].sum()),
    }

    # Resolve the client source branch (which is really Channel ID) to the true
    # parent branch, group, channel type, and channel name.
    map_df = ch[["Channel ID", "Branch Code", "Channel Type", "Group", "Channel Name", "State"]].drop_duplicates("Channel ID")
    e = cl.rename(columns={"Branch Code": "Channel ID"}).merge(map_df, on="Channel ID", how="left")
    e["Channel Name"] = e["Channel Name"].fillna(e["Branch Name"])
    e["Branch Code"] = e["Branch Code"].fillna("")
    e["Group"] = e["Group"].fillna("")
    e["Channel Type"] = e["Channel Type"].fillna("")
    e["State"] = e["State"].fillna("")

    # Collapse a client code/name across channels and select the primary channel by brokerage.
    client_key = ["Client Code", "Client Name"]
    sums = e.groupby(client_key, as_index=False)[["Turnover", "Gross Brokerage"]].sum()
    e_sorted = e.sort_values(client_key + ["Gross Brokerage", "Turnover", "Channel ID"], ascending=[True, True, False, False, True])
    primary = e_sorted.groupby(client_key, as_index=False).first()
    primary = primary.drop(columns=["Turnover", "Gross Brokerage"], errors="ignore").merge(sums, on=client_key, how="left")
    primary["Yield"] = (primary["Gross Brokerage"] / primary["Turnover"].replace(0, math.nan) * 100).fillna(0)
    primary["Is House Account"] = primary["Client Code"].isin(HOUSE_CODES)
    primary = primary.rename(columns={"Branch Code": "Parent Branch Code"})
    primary = primary[["Parent Branch Code", "Client Code", "Client Name", "City", "Channel ID", "Channel Name", "Group", "Channel Type", "State", "Turnover", "Gross Brokerage", "Yield", "Is House Account"]]
    primary = primary.sort_values(["Parent Branch Code", "Client Code", "Client Name"], kind="stable").reset_index(drop=True)

    gc.collect()
    return ch, primary, source_ch, source_cl


def pct_gap(a, b):
    return (a - b) / b * 100 if b else 0.0


def channel_rows(df):
    out = []
    for r in df.sort_values(["Branch Code", "Channel ID", "Channel Name"]).to_dict("records"):
        out.append([
            r["Branch Code"], r["Channel ID"], r["Channel Type"], r["Group"], r["Channel Name"], r["City"], r["State"],
            float(r["Turnover"] / 1e7), float(r["Gross Brokerage"] / 1e5), float(r["Passout"] / 1e5), float(r["Net Brokerage"] / 1e5), float(r["Yield"]), "Detail"
        ])
    return out


def client_rows(df, include_house=True):
    if not include_house:
        df = df[~df["Is House Account"]].copy()
    out = []
    for r in df.sort_values(["Parent Branch Code", "Client Code", "Client Name"]).to_dict("records"):
        out.append([
            r["Parent Branch Code"], r["Client Code"], r["Client Name"], r["City"], r["Channel ID"], r["Channel Name"], r["Group"], r["Channel Type"],
            bool(r["Is House Account"]), float(r["Turnover"] / 1e7), float(r["Gross Brokerage"] / 1e5), float(r["Yield"]), "Detail"
        ])
    return out


def with_subtotals(rows, branch_idx, metric_start_idx, yield_idx, label_idx, label_prefix):
    if not rows:
        return []
    out = []
    i = 0
    while i < len(rows):
        branch = rows[i][branch_idx]
        j = i
        while j < len(rows) and rows[j][branch_idx] == branch:
            j += 1
        details = rows[i:j]
        out.extend(details)
        total = list(details[0])
        # A subtotal row intentionally carries a blank Group-like field via the caller's layout.
        for k in range(len(total)):
            total[k] = "" if isinstance(total[k], str) else 0
        total[branch_idx] = ""
        total[label_idx] = f"{label_prefix} {branch}"
        for k in range(metric_start_idx, yield_idx + 1):
            if k == yield_idx:
                continue
            total[k] = sum(float(r[k] or 0) for r in details)
        turnover = total[metric_start_idx]
        gross = total[metric_start_idx + 1] if yield_idx > metric_start_idx + 1 else 0
        # Channel rows: metric_start is turnover, gross is next; client rows same.
        total[yield_idx] = (gross / turnover * 100) if turnover else 0
        total[-1] = "Branch Total"
        out.append(total)
        i = j
    return out


def channel_map_by_id(df):
    return {r["Channel ID"]: r for r in df.to_dict("records")}


def make_channel_comparison(fy, new, factor):
    fy_map = channel_map_by_id(fy)
    new_map = channel_map_by_id(new)
    rows = []
    for key in sorted(set(fy_map) | set(new_map)):
        a = fy_map.get(key, {})
        b = new_map.get(key, {})
        turnover_fy = float(a.get("Turnover", 0))
        turnover_new = float(b.get("Turnover", 0))
        gross_fy = float(a.get("Gross Brokerage", 0))
        gross_new = float(b.get("Gross Brokerage", 0))
        pass_fy = float(a.get("Passout", 0))
        pass_new = float(b.get("Passout", 0))
        net_fy = float(a.get("Net Brokerage", 0))
        net_new = float(b.get("Net Brokerage", 0))
        fy_y = gross_fy / turnover_fy * 100 if turnover_fy else 0
        new_y = gross_new / turnover_new * 100 if turnover_new else 0
        rows.append([
            a.get("Branch Code", b.get("Branch Code", "")), key, a.get("Channel Type", b.get("Channel Type", "")), a.get("Group", b.get("Group", "")),
            a.get("Channel Name", b.get("Channel Name", "")), a.get("City", b.get("City", "")), a.get("State", b.get("State", "")),
            turnover_fy / 1e7, gross_fy / 1e5, pass_fy / 1e5, net_fy / 1e5, fy_y,
            turnover_new / 1e7, gross_new / 1e5, pass_new / 1e5, net_new / 1e5, new_y,
            factor, turnover_new * factor / 1e7, gross_new * factor / 1e5, pass_new * factor / 1e5, net_new * factor / 1e5, new_y,
            ((gross_new * factor / gross_fy) - 1) if gross_fy else None,
            ((pass_new * factor / pass_fy) - 1) if pass_fy else None,
            ((net_new * factor / net_fy) - 1) if net_fy else None,
            ((new_y / fy_y) - 1) if fy_y else None,
            "New" if not a else ("Lost" if not b else "Matched")
        ])
    return rows


def make_client_comparison(fy, new, factor):
    keys = ["Client Code", "Client Name"]
    keep = ["Parent Branch Code", "City", "Channel ID", "Channel Name", "Group", "Channel Type", "Turnover", "Gross Brokerage", "Yield"]
    fy2 = fy[~fy["Is House Account"]][keys + keep].rename(columns={c: f"{c} FY" for c in keep})
    new2 = new[~new["Is House Account"]][keys + keep].rename(columns={c: f"{c} New" for c in keep})
    merged = fy2.merge(new2, on=keys, how="outer", indicator=True)
    matched, newly, lost = [], [], []
    for d in merged.sort_values(keys).to_dict("records"):
        status = d["_merge"]
        a_exists = status in ("both", "left_only")
        b_exists = status in ("both", "right_only")
        ft = float(d.get("Turnover FY", 0) or 0) if a_exists else 0
        nt = float(d.get("Turnover New", 0) or 0) if b_exists else 0
        fg = float(d.get("Gross Brokerage FY", 0) or 0) if a_exists else 0
        ng = float(d.get("Gross Brokerage New", 0) or 0) if b_exists else 0
        fy_y = float(d.get("Yield FY", 0) or 0) if a_exists else 0
        new_y = float(d.get("Yield New", 0) or 0) if b_exists else 0
        base_prefix = "New" if b_exists else "FY"
        def base_val(name):
            return d.get(f"{name} {base_prefix}", "")
        row = [
            base_val("Parent Branch Code"), d["Client Code"], d["Client Name"], base_val("City"), base_val("Channel ID"), base_val("Channel Name"), base_val("Group"), base_val("Channel Type"),
            ft / 1e7, fg / 1e5, fy_y, nt / 1e7, ng / 1e5, new_y, factor, nt * factor / 1e7, ng * factor / 1e5, new_y,
            (ng * factor / fg - 1) if fg else None, (ng * factor - fg) / 1e5, "Matched" if a_exists and b_exists else ("New" if b_exists else "Lost")
        ]
        if a_exists and b_exists:
            matched.append(row)
        elif b_exists:
            newly.append(row)
        else:
            lost.append(row)
    return matched, newly, lost


def top_rows(rows, n):
    return sorted(rows, key=lambda r: float(r[16] or 0), reverse=True)[:n]


def gainers_decliners(matched, n=20):
    valid = [r for r in matched if r[19] is not None]
    return sorted(valid, key=lambda r: float(r[19]), reverse=True)[:n], sorted(valid, key=lambda r: float(r[19]))[:n]


import re

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def parse_period_from_filename(filename):
    """Extract date range and labels from a filename like 'Opr MIS Data Apr25 to Mar26.xlsx'."""
    m = re.search(r"([A-Za-z]{3})(\d{2})\s*to\s*([A-Za-z]{3})(\d{2})", filename)
    if not m:
        return None
    start_mon, start_yr, end_mon, end_yr = m.group(1).lower(), int(m.group(2)), m.group(3).lower(), int(m.group(4))
    start_month_num = MONTH_MAP.get(start_mon, 4)
    end_month_num = MONTH_MAP.get(end_mon, 3)
    # Calculate months covered
    start_abs = (2000 + start_yr) * 12 + start_month_num
    end_abs = (2000 + end_yr) * 12 + end_month_num
    months = end_abs - start_abs + 1
    # Fiscal year: Apr starts a new FY
    fy_start_year = 2000 + start_yr if start_month_num >= 4 else 2000 + start_yr - 1
    fy_label = f"FY {fy_start_year}-{str(fy_start_year + 1)[-2:]}"
    short_label = f"{m.group(1).title()}'{start_yr:02d}–{m.group(3).title()}'{end_yr:02d} Actual"
    long_label = f"{m.group(1).title()}'{start_yr:02d}–{m.group(3).title()}'{end_yr:02d} ({fy_label}, {months} months actual)"
    is_full_year = months == 12
    return {
        "months": months,
        "factor": 12 // months if months < 12 else 1,
        "fyLabel": fy_label,
        "shortLabel": short_label,
        "longLabel": long_label,
        "isFullYear": is_full_year,
    }


def process(fy_path=None, new_path=None):
    """Run the full preprocessing pipeline. Returns (payload_dict, stats_dict).
    If paths are not given, uses the default hardcoded files."""
    fy_path = Path(fy_path) if fy_path else FY_FILE
    new_path = Path(new_path) if new_path else NEW_FILE

    # Parse period info from filenames
    fy_info = parse_period_from_filename(fy_path.name) or {
        "months": 12, "factor": 1, "fyLabel": "FY 2025-26",
        "shortLabel": "FY 2025-26", "longLabel": "FY 2025-26 (12 months actual)", "isFullYear": True,
    }
    new_info = parse_period_from_filename(new_path.name) or {
        "months": 4, "factor": 3, "fyLabel": "FY 2026-27",
        "shortLabel": "New Period Actual", "longLabel": "New Period", "isFullYear": False,
    }
    factor = new_info["factor"]

    fy_ch, fy_cl, fy_ch_src, fy_cl_src = read_source(fy_path)
    new_ch, new_cl, new_ch_src, new_cl_src = read_source(new_path)

    fy_channel_detail = channel_rows(fy_ch)
    new_channel_detail = channel_rows(new_ch)
    fy_client_detail = client_rows(fy_cl, include_house=True)
    new_client_detail = client_rows(new_cl, include_house=True)
    channel_comp = make_channel_comparison(fy_ch, new_ch, factor)
    matched, new_clients, lost_clients = make_client_comparison(fy_cl, new_cl, factor)
    all_clients = sorted(matched + new_clients + lost_clients, key=lambda r: float(r[16] or 0), reverse=True)
    top100 = all_clients[:100]
    top500 = all_clients[:500]
    gainers, decliners = gainers_decliners(matched)

    # Confirm groups present in data
    actual_groups = sorted(set(fy_ch["Group"].unique()) | set(new_ch["Group"].unique()))
    actual_groups = [g for g in actual_groups if g]  # remove blanks
    groups = actual_groups if actual_groups else GROUPS

    def summarize_clients(df):
        x = df[~df["Is House Account"]].groupby("Group", as_index=False).agg(
            Clients=("Client Code", "count"), Turnover=("Turnover", "sum"), Gross=("Gross Brokerage", "sum")
        )
        x["Yield"] = (x["Gross"] / x["Turnover"].replace(0, math.nan) * 100).fillna(0)
        return [[r["Group"], int(r["Clients"]), float(r["Turnover"] / 1e7), float(r["Gross"] / 1e5), float(r["Yield"])] for r in x.to_dict("records")]

    def summarize_channels(df):
        x = df.groupby("Group", as_index=False).agg(
            Channels=("Channel ID", "count"), Turnover=("Turnover", "sum"), Gross=("Gross Brokerage", "sum"), Passout=("Passout", "sum"), Net=("Net Brokerage", "sum")
        )
        x["Yield"] = (x["Gross"] / x["Turnover"].replace(0, math.nan) * 100).fillna(0)
        return [[r["Group"], int(r["Channels"]), float(r["Turnover"] / 1e7), float(r["Gross"] / 1e5), float(r["Passout"] / 1e5), float(r["Net"] / 1e5), float(r["Yield"])] for r in x.to_dict("records")]

    recon = {
        fy_info["fyLabel"]: {
            "channelwise": fy_ch_src, "clientwise": fy_cl_src,
            "turnover_gap_pct": pct_gap(fy_ch_src["turnover"], fy_cl_src["turnover"]),
            "gross_gap_pct": pct_gap(fy_ch_src["gross"], fy_cl_src["gross"]),
        },
        new_info["longLabel"]: {
            "channelwise": new_ch_src, "clientwise": new_cl_src,
            "turnover_gap_pct": pct_gap(new_ch_src["turnover"], new_cl_src["turnover"]),
            "gross_gap_pct": pct_gap(new_ch_src["gross"], new_cl_src["gross"]),
        },
    }

    payload = {
        "meta": {
            "fyLabel": fy_info["fyLabel"],
            "newLabel": new_info["longLabel"],
            "newShortLabel": new_info["shortLabel"],
            "factor": factor,
            "groups": groups,
            "houseCodes": sorted(HOUSE_CODES),
            "reconciliation": recon,
        },
        "schemas": {
            "channelActual": ["Branch Code", "Channel ID", "Channel Type", "Group", "Channel Name", "City", "State", "Turnover (₹ Crore) [helper]", "Gross Brokerage (₹ Lacs)", "Passout (₹ Lacs)", "Net Brokerage (₹ Lacs)", "Yield %", "Row Type"],
            "clientActual": ["Branch Code", "Client Code", "Client Name", "City", "Channel ID", "Channel Name", "Group", "Channel Type", "Is House Account", "Turnover (₹ Crore)", "Gross Brokerage (₹ Lacs)", "Yield %", "Row Type"],
            "channelComparison": ["Branch Code", "Channel ID", "Channel Type", "Group", "Channel Name", "City", "State", "FY Turnover (₹ Crore) [helper]", "FY Gross (₹ Lacs)", "FY Passout (₹ Lacs)", "FY Net (₹ Lacs)", "FY Yield %", "New Turnover Actual (₹ Crore) [helper]", "New Gross Actual (₹ Lacs)", "New Passout Actual (₹ Lacs)", "New Net Actual (₹ Lacs)", "New Yield Actual %", "Annualization Factor", "Annualized New Turnover (₹ Crore) [helper]", "Annualized New Gross (₹ Lacs)", "Annualized New Passout (₹ Lacs)", "Annualized New Net (₹ Lacs)", "Annualized New Yield %", "Gross Growth %", "Passout Growth %", "Net Growth %", "Yield Growth %", "Status"],
            "clientComparison": ["Branch Code", "Client Code", "Client Name", "City", "Channel ID", "Channel Name", "Group", "Channel Type", "FY Turnover (₹ Crore)", "FY Gross (₹ Lacs)", "FY Yield %", "New Turnover Actual (₹ Crore)", "New Gross Actual (₹ Lacs)", "New Yield Actual %", "Annualization Factor", "Annualized New Turnover (₹ Crore)", "Annualized New Gross (₹ Lacs)", "Annualized New Yield %", "Gross Growth %", "Absolute Gross Change (₹ Lacs)", "Status"],
        },
        "channels": {
            "fyDetails": fy_channel_detail,
            "newDetails": new_channel_detail,
            "comparison": channel_comp,
            "groupFY": summarize_channels(fy_ch),
            "groupNew": summarize_channels(new_ch),
        },
        "clients": {
            "fyDetails": fy_client_detail,
            "newDetails": new_client_detail,
            "groupFY": summarize_clients(fy_cl),
            "groupNew": summarize_clients(new_cl),
            "matched": matched,
            "new": new_clients,
            "lost": lost_clients,
            "top100": top100,
            "top500": top500,
            "gainers": gainers,
            "decliners": decliners,
        },
    }

    # Save to build dir
    (OUT / "data.json").write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    stats = {
        "fy_channels": len(fy_channel_detail), "new_channels": len(new_channel_detail),
        "fy_clients": len(fy_client_detail), "new_clients": len(new_client_detail),
        "matched_clients": len(matched), "new_clients_comparison": len(new_clients), "lost_clients": len(lost_clients),
        "recon": recon,
    }
    return payload, stats


def main():
    payload, stats = process()
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
